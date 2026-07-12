"""One-shot: convert the class-roster .xlsx workbooks into the import CSV.

Dev-only. `openpyxl` is a [dev] dependency and is deliberately NOT in the
runtime image — the production container never reads these files.

    pip install -e ".[dev]"
    python scripts/convert_class_rosters.py

Reads   private-data/*.xlsx   (gitignored — REAL NAMES, this repo is public)
Writes  private-data/class_rosters.csv  (gitignored)

Then:   python manage.py import_class_roster private-data/class_rosters.csv --dry-run

## Why the column order is hardcoded rather than read from the header

The two workbooks disagree: `6ème promotion 80 - 81.xlsx` labels its columns
"Noms | Prénoms", while `6ème 81 - 82.xlsx` labels the same physical columns
"Prénoms | Noms". The data, not the headers, settles it:

  * all 12 people who appear in BOTH workbooks have their name parts in the
    same physical order in each, and
  * the owner's own row reads `Mahamadou | Laouali` with surnom "Bomino".

So column B is the GIVEN name and column C the FAMILY name in both files, and
the 80-81 header is simply mislabeled. Trust the data.
"""

from __future__ import annotations

import csv
import re
import sys
import unicodedata
from pathlib import Path

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:  # pragma: no cover - dev-only script
    sys.exit("openpyxl is required: pip install -e '.[dev]'")

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "private-data"
OUT_CSV = DATA_DIR / "class_rosters.csv"

# (workbook filename, tag used in source_ref, school year the 6ème started)
SOURCES = [
    ("6ème promotion 80 - 81.xlsx", "80-81", 1980),
    ("6ème 81 - 82.xlsx", "81-82", 1981),
]

FIELDNAMES = [
    "source_ref",
    "school_year_start",
    "class_label",
    "first_name",
    "last_name",
    "nickname",
    "needs_review",
]


def _clean(value) -> str:
    """Trim and collapse internal whitespace. Accents are preserved."""
    if value is None:
        return ""
    return " ".join(str(value).split())


def _class_label(sheet_name: str) -> str:
    """'6ème A' -> '6eA'. Matches members.models.VALID_CLASS_PATTERN."""
    m = re.search(r"([3-6])\s*(?:e|è)me\s*([A-Za-z])?", sheet_name, flags=re.IGNORECASE)
    if not m:
        raise ValueError(f"Cannot derive a class label from sheet {sheet_name!r}")
    grade, section = m.group(1), (m.group(2) or "").upper()
    return f"{grade}e{section}"


def _norm(text: str) -> str:
    """Accent- and case-insensitive form, for duplicate detection only."""
    folded = unicodedata.normalize("NFD", text.casefold())
    return "".join(c for c in folded if unicodedata.category(c) != "Mn")


def _token_key(first: str, last: str) -> str:
    """Order-insensitive identity key, so 'Aledji Viviane' == 'Viviane Aledji'."""
    return " ".join(sorted(f"{_norm(first)} {_norm(last)}".split()))


def main() -> int:
    rows: list[dict] = []
    seen: dict[str, str] = {}  # token key -> first source_ref that used it

    for filename, tag, year in SOURCES:
        path = DATA_DIR / filename
        if not path.exists():
            sys.exit(f"Missing {path}. Keep the workbooks in private-data/ (gitignored).")

        wb = load_workbook(path, data_only=True)
        for sheet in wb.worksheets:
            label = _class_label(sheet.title)
            for row_no, raw in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                cells = [_clean(c) for c in raw]
                # Column B = given name, C = family name, D = surnom (see docstring).
                first = cells[1] if len(cells) > 1 else ""
                last = cells[2] if len(cells) > 2 else ""
                nickname = cells[3] if len(cells) > 3 else ""

                if not first and not last:
                    continue  # numbered filler row

                # A person with no family name in the source is legitimate (20 of them);
                # keep the row, leave last_name blank.
                key = _token_key(first, last)
                needs_review = False
                if not last:
                    needs_review = True  # single-name row — owner should complete it
                if key in seen:
                    # Same person listed twice. Across workbooks that's a repeated 6ème
                    # (legitimate); within one workbook it's a data error. Either way the
                    # owner should look at it.
                    needs_review = True

                source_ref = f"{tag}:{label}:{row_no}"
                seen.setdefault(key, source_ref)

                rows.append(
                    {
                        "source_ref": source_ref,
                        "school_year_start": year,
                        "class_label": label,
                        "first_name": first,
                        "last_name": last,
                        "nickname": nickname,
                        "needs_review": "1" if needs_review else "",
                    }
                )

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with OUT_CSV.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=FIELDNAMES)
        writer.writeheader()
        writer.writerows(rows)

    classes = sorted({(r["school_year_start"], r["class_label"]) for r in rows})
    flagged = sum(1 for r in rows if r["needs_review"])
    print(f"Wrote {OUT_CSV} — {len(rows)} entries across {len(classes)} classes.")
    for year, label in classes:
        n = sum(1 for r in rows if r["school_year_start"] == year and r["class_label"] == label)
        print(f"  {year}-{year + 1}  {label}: {n}")
    print(f"\n{flagged} row(s) flagged needs_review (blank surname or listed twice).")
    print("Review them in /admin/members/classrosterentry/ after importing.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
